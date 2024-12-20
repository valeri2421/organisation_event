#Модуль с методами для работы с БД.
import sqlite3 as sq
from datetime import datetime
from openpyxl import load_workbook

db = sq.connect('system_bd.db')
cur = db.cursor()

def add_administration (user_id, pin):
    cur.execute("UPDATE administration SET user_ID = ? WHERE password = ?",
                (user_id, pin))
    db.commit()

def add_organizer (user_id, pin):
    cur.execute("UPDATE organizers SET user_ID = ? WHERE password_org = ?",
                (user_id, pin))
    db.commit()

def insert_event(name, type, place, date_time_start, date_time_end, status, kol_org, estimate):
    cur.execute("""INSERT INTO events (name, type, place, date_time_start, date_time_end, status, kol_org, estimate) 
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (name, type, place, date_time_start, date_time_end, status, kol_org, estimate))

    db.commit()

def delete_admin_id(user_id):
    cur.execute("UPDATE administration SET user_ID = NULL WHERE user_ID = ?", (user_id,))
    db.commit()

def delete_org_id(user_id):
    cur.execute("UPDATE organizers SET user_ID = NULL WHERE user_ID = ?", (user_id,))
    db.commit()
def ifevent(name, date_time_start):
    cur.execute("""SELECT COUNT(*) FROM events WHERE name = ? AND date_time_start = ?""",
                (name, date_time_start))

def add_event(file_path):
# -----добавление мероприятия из екселя в бд-----
    wb = load_workbook(file_path)
    sheet = wb.active
    name = sheet['B3'].value
    type = sheet['B9'].value
    place = sheet['B4'].value
    date_time_start = datetime.strptime(sheet['B5'].value + ' ' + sheet['B6'].value, "%d.%m.%Y %H:%M")
    date_time_end = datetime.strptime(sheet['B7'].value + ' ' + sheet['B8'].value, "%d.%m.%Y %H:%M")
    date = datetime.now()
    if date < date_time_start:
        status = 'process'
    else:
        status = 'end'
    kol_org = 2
    if sheet['B12'].value == 'да' and type == 'гибрид': #Если есть трансляция, то к ним идет еще 1 организатор
        kol_org += 1
    estimate = 'нет'
    # Проверка на наличие мероприятия в БД
    k = cur.execute("""SELECT COUNT(*) FROM events WHERE name = ? AND date_time_start = ?""",
                    (name, date_time_start)).fetchone()[0]
    if k > 0:
        raise Exception('Мероприятие уже было добавлено')

    insert_event(name, type, place, date_time_start, date_time_end, status, kol_org, estimate)

def bd_req(list, num): # Запрос на проверку наличия сметы
    name = list[num][1]
    date, res = cur.execute(f"SELECT date_time_start, estimate FROM events WHERE name LIKE '%{name}%'").fetchone()
    return [res, name, date[:10]]


#Составление сметы
#video estimate
def video(name_event):
    wb_tz = load_workbook(name_event[1])
    sheet_tz = wb_tz.active
    res = ''
    if sheet_tz['B12'].value == 'да':
        res += 'Оборудование для видео отдела' + '\n'
        res += '\n'
        res += 'Комментарии от заказчика: %s' % (sheet_tz['D12'].value) + '\n'
        res += 'Справка: если есть экран, то + дежурный техник и виджей, если только проектор - видеоинженер' + '\n'
        res += '\n'
        alll = cur.execute("""SELECT * from equipment
                                            where sphere LIKE '%video%'""").fetchall()
        lis = []
        for i in range(len(alll)):
            res += f'{i + 1}.' + alll[i][2]+ ' ' + alll[i][3] + ' ' + str(alll[i][4]) + 'р.' + '\n'
            lis.append(alll[i])
        return [res, lis]
    else:
        return [res]


def other(name_event):
    res = ''
    res += 'Коммутационное оборудование' + '\n'
    res += 'Справка: кабель каналы -минимум 5 штук, минимум 10 штук xlr, dmx, 2 hdmi (далее смотреть по надобности из тз)'
    res += '\n'
    alll = cur.execute("""SELECT * from equipment
                                            where sphere LIKE '%other%'""").fetchall()
    lis = []
    for i in range(len(alll)):
        res += f'{i + 1}.' + alll[i][2]+ ' ' + alll[i][3] + ' ' + str(alll[i][4]) + 'р.' + '\n'
        lis.append(alll[i])
    return [res, lis]


def light(name_event):
    wb_tz = load_workbook(name_event[1])
    sheet_tz = wb_tz.active
    res = ''
    if sheet_tz['B15'].value == 'да':
        res += 'Оборудование для светового отдела' + '\n'
        res += '\n'
        res += 'Комментарии от заказчика: %s' % (sheet_tz['D15'].value) + '\n'
        res += '\n'
        alll = cur.execute("""SELECT * from equipment
                                            where sphere LIKE '%light%'""").fetchall()
        lis = []
        for i in range(len(alll)):
            res += f'{i + 1}.' + alll[i][2]+ ' ' + alll[i][3] + ' ' + str(alll[i][4]) + 'р.' + '\n'
            lis.append(alll[i])
        return [res, lis]
    else:
        return [res]


def sound(name_event):
    wb_tz = load_workbook(name_event[1])
    sheet_tz = wb_tz.active
    res = ''
    play = ""
    lis = []
    if sheet_tz['B14'].value == 'да':
        res += 'Оборудование для звукового отдела' + '\n'
        res += '\n'
        res += 'Комментарии от заказчика: %s' % (sheet_tz['D14'].value) + '\n'
        res += 'Справка: для каждого вокалиста по 1 микрофону, у барабанщика минимум 3 ударных, скрипки/трубы - бодипаки, при добавлениии людей - считать плейбекера за звукорежиссера' + '\n'
        res += '\n'
        alll = cur.execute("""SELECT * from equipment
                                            where sphere LIKE '%sound%'""").fetchall()

        for i in range(len(alll)):
            res += f'{i + 1}.' + alll[i][2]+ ' ' + alll[i][3] + ' ' + str(alll[i][4]) + 'р.' + '\n'
            lis.append(alll[i])
    if sheet_tz['B10'].value == 'да':
        new_file_path = name_event[2]
        workbook_sm = load_workbook(new_file_path)
        sheet_sm = workbook_sm.active
        if sheet_sm['H3'].value is None:
            list_desk = cur.execute("""SELECT * from equipment
                                                where affiliation LIKE '%Ноутбук%'""").fetchall()
            sheet_sm['H3'] = list_desk[0][2]
            sheet_sm['I3'] = list_desk[0][3]
            sheet_sm['J3'] = 1
            sheet_sm['K3'] = list_desk[0][4]
            sheet_sm['L3'] = 1
            sheet_sm['C41'] = 1
            play = "Ноутбук для плейбекера мероприятия добавлен в смету по тз"
            workbook_sm.save(new_file_path)
    return [res, lis, play]

def add_eq_otdel(num, name_event, eq):
    new_file_path = name_event[2]
    workbook_sm = load_workbook(new_file_path)
    sheet_sm = workbook_sm.active
    #нужно очищать потом список name_event, чтобы удалять оборудования с конца каждого отдела и чистить потом после всего добавления сметы этот список
    i = int(num[1:])
    k = sheet_sm[f"{num[0]}{i}"].value
    while not (k is None):
        i += 1
        k = sheet_sm[f"{num[0]}{i}"].value
    #print(name_event)
    try:
        n_eq, kol = eq.split(', ')
        n_eq, kol = int(n_eq) - 1, int(kol)
        n = name_event[3][n_eq]
        sheet_sm[f"{num[0]}{i}"] = n[2]
        next_ = chr(ord(num[0]) + 1)
        sheet_sm[f"{next_}{i}"] = n[3]
        next_ = chr(ord(next_) + 1)
        sheet_sm[f"{next_}{i}"] = kol
        next_ = chr(ord(next_) + 1)
        sheet_sm[f"{next_}{i}"] = n[4]
        next_ = chr(ord(next_) + 1)
        sheet_sm[f"{next_}{i}"] = 1
        workbook_sm.save(new_file_path)
        return [name_event[:-2], True]
    except:
        return [name_event[:-2], False]


def add_people_smeta(n, name_event):
    new_file_path = name_event[2]
    workbook_sm = load_workbook(new_file_path)
    sheet_sm = workbook_sm.active
    duty, kol_people, taxi = 0, 0, 0
    try:
        n = n.split(', ')
        for i in n:
            k = i.split(' - ')
            if 'идео' in k[0]:
                y = int(k[1])
                if not (sheet_sm['C40'].value is None):
                    y += sheet_sm['C40'].value
                sheet_sm['C40'] = y
                duty += 1 if y == 1 else y - 1
                kol_people += y
            elif 'удожник' in k[0]:
                y = int(k[1])
                if not (sheet_sm['C42'].value is None):
                    y += sheet_sm['C42'].value
                sheet_sm['C42'] = y
                duty += 1 if y == 1 else y - 1
                kol_people += y
            elif 'вукореж' in k[0]:
                y = int(k[1])
                if not (sheet_sm['C41'].value is None):
                    y += sheet_sm['C41'].value
                sheet_sm['C41'] = y
                duty += 1 if y == 1 else y - 1
                kol_people += y
        sheet_sm['C43'] = duty # автоматический расчет дежурных площадки
        #расчет транспорта
        new_file_path2 = name_event[1]
        workbook_tz = load_workbook(new_file_path2)
        sheet_tz = workbook_tz.active
        time_start = sheet_tz['B6'].value
        time_end = sheet_tz['B8'].value
        flag1, flag2 = False, False
        if int(time_start[:2]) < 6 and int(time_start[:2]) > 1:
            flag1 = True
        if int(time_end[:2]) > 1 and int(time_end[:2]) < 10:
            flag2 = True
        if flag1:
            taxi = kol_people * 500
        if flag2:
            taxi += kol_people * 500
        sheet_sm['D44'] = taxi
        workbook_sm.save(new_file_path)
        return True
    except:
        return False

def status_smeta(lis):
    new_file_path = lis[2]
    workbook_sm = load_workbook(new_file_path)
    sheet_sm = workbook_sm.active
    if sheet_sm['A3'].value is None and sheet_sm['H3'].value is None and sheet_sm['H20'].value is None and sheet_sm['A20'].value is None:
        return False
    else:
        return True
